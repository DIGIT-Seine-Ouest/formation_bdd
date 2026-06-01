#!/usr/bin/env python3
"""
Génère reclamations_complet.xlsx — MODÈLE DE CORRECTION du Serious Game.

Aligné EXACTEMENT sur les slides du Serious Game (#/27 et suivantes) :
  OLTP  : prestataires · lignes_bus · passe_par · ref_communes · ref_motifs · reclamations
  OLAP  : 📊 Tableau de bord  → répond aux 3 grandes questions métier.

Colonnes de [reclamations] calées sur les formules montrées dans les slides :
  A id_reclamation · B date · C id_ligne · D id_motif · E id_commune
  F objet_libre · G date_cloture · H cloturee · I delai_traitement (helper)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from datetime import date

# ── Palette ───────────────────────────────────────────────────────────────────
DARK   = "1E3A5F"   # OLTP central
BLUE   = "009FE3"
TEAL   = "0F766E"   # jonction
GREEN  = "15803D"
OLIVE  = "95C11F"   # OLAP
PURPLE = "7C3AED"
RED    = "C0392B"
AMBER  = "D97706"
SLATE  = "64748B"
GREY   = "94A3B8"

LT_BLUE, LT_GREEN, LT_PURPLE, LT_RED, LT_AMBER = "EFF6FF","F0FDF4","F5F0FF","FFF0F0","FFFBEB"
LT_TEAL, LT_OLIVE = "F0FDFA","F7FEE7"
WHITE, GRAY1, GRAY2 = "FFFFFF","F8FAFC","F1F5F9"

THIN = Side(style='thin', color="D1D5DB")
MED  = Side(style='medium', color=DARK)

# ── Helpers ─────────────────────────────────────────────────────────────────--
def C(n): return get_column_letter(n)

def cell(ws, r, c, v, *, bold=False, italic=False, color="222222", size=10,
         bg=None, align='left', valign='center', wrap=False, fmt=None, name='Calibri'):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(bold=bold, italic=italic, color=color, size=size, name=name)
    if bg: x.fill = PatternFill("solid", fgColor=bg)
    x.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if fmt: x.number_format = fmt
    return x

def header(ws, r, c, txt, bg=DARK, fg=WHITE, size=10):
    return cell(ws, r, c, txt, bold=True, color=fg, bg=bg, align='center', wrap=True, size=size)

def box(ws, r1, r2, c1, c2, side=THIN):
    for row in ws.iter_rows(r1, r2, c1, c2):
        for x in row:
            x.border = Border(left=side, right=side, top=side, bottom=side)

def stripe(i, even=LT_BLUE, odd=WHITE): return even if i % 2 == 0 else odd

def setup(ws):
    ws.sheet_view.showGridLines = False

# ════════════════════════════════════════════════════════════════════════════
#  DONNÉES  (calées sur les slides ; "coupable" net : Transdev / Meudon / L169)
# ════════════════════════════════════════════════════════════════════════════
PRESTATAIRES = [
    # id    nom                       type_contrat  debut        fin
    ("P001","Keolis Île-de-France",   "DSP",        date(2022,1,1),  date(2026,12,31)),
    ("P002","RATP Dev",               "DSP",        date(2020,6,1),  date(2025,5,31)),
    ("P003","Transdev Île-de-France", "DSP",        date(2021,3,1),  date(2026,2,28)),
    ("P004","ATM Croix du Sud",       "Convention", date(2023,1,1),  date(2027,12,31)),
]

LIGNES = [
    # id_ligne  numero  id_prestataire  type_service  actif
    ("L169",  "169",  "P003", "local",    1),
    ("L323",  "323",  "P003", "local",    1),
    ("L389",  "389",  "P004", "local",    1),
    ("L290",  "290",  "P004", "local",    0),   # ligne suspendue → montre le flag actif
    ("L058",  "58",   "P002", "local",    1),
    ("LN118", "N118", "P002", "régional", 1),
    ("L291",  "291",  "P001", "local",    1),
    ("L289",  "289",  "P001", "local",    1),
]

COMMUNES = [
    ("C001","Boulogne-Billancourt"),
    ("C002","Issy-les-Moulineaux"),
    ("C003","Meudon"),
    ("C004","Clamart"),
    ("C005","Chaville"),
    ("C006","Sèvres"),
    ("C007","Vanves"),
    ("C008","Ville-d'Avray"),
]

# Table de jonction : une ligne dessert plusieurs communes (many-to-many)
PASSE_PAR = [
    ("L169","C004"),("L169","C003"),("L169","C008"),
    ("L323","C002"),("L323","C005"),("L323","C006"),
    ("L389","C003"),("L389","C002"),("L389","C001"),
    ("L290","C001"),("L290","C004"),
    ("L058","C001"),("L058","C002"),("L058","C007"),
    ("LN118","C003"),("LN118","C005"),("LN118","C008"),
    ("L291","C005"),("L291","C006"),
    ("L289","C004"),("L289","C007"),
]

MOTIFS = [
    (1,"Retard"),
    (2,"Information voyageurs"),
    (3,"Suppression de service"),
    (4,"Comportement conducteur"),
    (5,"Confort / Surcharge"),
    (6,"Infrastructure / Arrêt"),
]

# id  date         ligne   motif commune objet                                            cloture     clos
RECLAMATIONS = [
    ("R001",date(2024,1,5), "L169", 1,"C003","Retard de 20 min à l'arrêt Meudon-Val Fleury", date(2024,1,7), 1),
    ("R002",date(2024,1,9), "L169", 1,"C004","Retard répété tous les matins",                None,           0),
    ("R003",date(2024,1,12),"L323", 3,"C003","Bus supprimé sans annonce préalable",          None,           0),
    ("R004",date(2024,1,18),"L169", 1,"C003","Retard systématique vers 8h",                  None,           0),
    ("R005",date(2024,1,22),"L058", 5,"C001","Bus bondé, impossible de monter",              date(2024,1,25),1),
    ("R006",date(2024,2,1), "L389", 4,"C002","Chauffeur impoli envers une passagère",        date(2024,2,4), 1),
    ("R007",date(2024,2,5), "L169", 2,"C008","Horaires incorrects sur l'application mobile", None,           0),
    ("R008",date(2024,2,9), "L323", 1,"C006","Retard de 30 min, correspondance ratée",       date(2024,2,11),1),
    ("R009",date(2024,2,14),"L169", 1,"C003","Retard récurrent en heure de pointe",          None,           0),
    ("R010",date(2024,2,19),"LN118",3,"C005","Suppression sans information aux voyageurs",   date(2024,2,22),1),
    ("R011",date(2024,2,25),"L291", 6,"C005","Abribus vandalisé, vitre cassée",              date(2024,3,2), 1),
    ("R012",date(2024,3,1), "L323", 1,"C002","Retard dû à des travaux non signalés",         None,           0),
    ("R013",date(2024,3,5), "L058", 4,"C007","Comportement agressif du conducteur",          date(2024,3,7), 1),
    ("R014",date(2024,3,8), "L169", 3,"C003","Bus supprimé deux fois dans la journée",       None,           0),
    ("R015",date(2024,3,12),"L289", 5,"C004","Surcharge dangereuse en heure de pointe",      date(2024,3,15),1),
    ("R016",date(2024,3,15),"L389", 1,"C001","Retard de 25 min sans information",            date(2024,3,17),1),
    ("R017",date(2024,3,19),"LN118",6,"C003","Arrêt non accessible aux personnes à mobilité réduite", None, 0),
    ("R018",date(2024,3,22),"L323", 2,"C006","Panneau d'affichage en panne",                 date(2024,3,25),1),
    ("R019",date(2024,3,26),"L169", 1,"C003","Retard matinal systématique",                  None,           0),
    ("R020",date(2024,4,2), "L290", 1,"C004","Bus non passé à l'heure prévue",               date(2024,4,4), 1),
    ("R021",date(2024,4,5), "L291", 5,"C006","Surcharge en cabine, sécurité compromise",     None,           0),
    ("R022",date(2024,4,9), "L289", 4,"C007","Échange verbal vif conducteur / passager",     date(2024,4,11),1),
]

PRESTA_OF_LIGNE = {l[0]: l[2] for l in LIGNES}
NAME_OF_PRESTA  = {p[0]: p[1] for p in PRESTATAIRES}
NAME_OF_COMMUNE = {c[0]: c[1] for c in COMMUNES}

# ════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ────────────────────────────────────────────────────────────────────────────
#  0. LEXIQUE / README
# ────────────────────────────────────────────────────────────────────────────
lx = wb.active
lx.title = "Lexique"
setup(lx)
lx.column_dimensions['A'].width = 2
lx.column_dimensions['B'].width = 30
lx.column_dimensions['C'].width = 62
lx.column_dimensions['D'].width = 30

lx.merge_cells('B1:D1')
cell(lx, 1, 2, "RÉCLAMATIONS USAGERS — MODÈLE DE CORRECTION COMPLET",
     bold=True, color=WHITE, bg=DARK, size=14, align='center')
lx.row_dimensions[1].height = 36
lx.merge_cells('B2:D2')
cell(lx, 2, 2, "Serious Game · de la construction des tables (OLTP) au tableau de suivi (OLAP)",
     italic=True, color=SLATE, size=10, align='center')

# OLTP block
cell(lx, 4, 2, "① PARTIE OLTP — saisir & stocker", bold=True, color=BLUE, size=12)
oltp_rows = [
    ("prestataires",   "Les opérateurs de transport — un contrat par prestataire."),
    ("lignes_bus",     "Les lignes. FK → prestataires. Flag actif pour suspendre sans supprimer."),
    ("passe_par",      "Table de JONCTION : une ligne dessert plusieurs communes (n–n)."),
    ("ref_communes",   "Référentiel des communes desservies."),
    ("ref_motifs",     "Liste fermée des motifs de réclamation (1 à 6)."),
    ("reclamations",   "TABLE CENTRALE — 1 ligne = 1 réclamation. FK vers ligne, motif, commune."),
]
r = 5
for nom, desc in oltp_rows:
    cell(lx, r, 2, nom, bold=True, color=DARK, bg=LT_BLUE, name='Consolas')
    cell(lx, r, 3, desc, color="333333", bg=LT_BLUE, wrap=True)
    r += 1

# OLAP block
r += 1
cell(lx, r, 2, "② PARTIE OLAP — lire & analyser", bold=True, color=GREEN, size=12); r += 1
cell(lx, r, 2, "📊 Tableau de bord", bold=True, color=GREEN, bg=LT_OLIVE, name='Consolas')
cell(lx, r, 3, "Vue de synthèse. Se recalcule seul à chaque nouvelle réclamation saisie.",
     color="333333", bg=LT_OLIVE, wrap=True); r += 2

# Les 3 questions
cell(lx, r, 2, "LES 3 GRANDES QUESTIONS auxquelles répond le tableau de bord :",
     bold=True, color=DARK, size=11); r += 1
for q in ["🔴  Quel est le prestataire le plus problématique ? (réclamations non clôturées)",
          "🟣  Quelle commune compte le plus de réclamations ?",
          "🟢  Quelle ligne pose le plus de problèmes ?"]:
    cell(lx, r, 2, q, color="333333", bg=GRAY1); lx.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4); r += 1

r += 1
cell(lx, r, 2, "🔗 Chaîne de jointure :", bold=True, color=TEAL); r += 1
cell(lx, r, 2, "reclamations → lignes_bus → prestataires    |    reclamations → ref_communes    |    lignes_bus ←→ passe_par ←→ ref_communes",
     italic=True, color=SLATE, name='Consolas', size=9)
lx.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

# ────────────────────────────────────────────────────────────────────────────
#  1. prestataires  (OLTP — Équipe A)
# ────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("prestataires"); setup(ws); ws.freeze_panes = "A2"
cols = ["id_prestataire","nom_prestataire","type_contrat","date_debut_contrat","date_fin_contrat"]
for i,h in enumerate(cols,1): header(ws,1,i,h)
ws.row_dimensions[1].height = 26
for r,(idp,nom,typ,deb,fin) in enumerate(PRESTATAIRES,2):
    bg = stripe(r)
    cell(ws,r,1,idp,bold=True,color=DARK,align='center',bg=bg,name='Consolas')
    cell(ws,r,2,nom,bg=bg)
    cell(ws,r,3,typ,align='center',bg=bg)
    cell(ws,r,4,deb,align='center',bg=bg,fmt='DD/MM/YYYY')
    cell(ws,r,5,fin,align='center',bg=bg,fmt='DD/MM/YYYY')
for c,w in zip("ABCDE",[16,26,14,18,18]): ws.column_dimensions[c].width = w
box(ws,1,len(PRESTATAIRES)+1,1,5)

# ────────────────────────────────────────────────────────────────────────────
#  2. lignes_bus  (OLTP — Équipe B)  + colonnes helper nb_reclam / nb_non_clot
# ────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("lignes_bus"); setup(ws); ws.freeze_panes = "A2"
cols = ["id_ligne","numero","id_prestataire","type_service","actif",
        "nb_reclam","nb_non_clot"]
for i,h in enumerate(cols,1):
    header(ws,1,i,h, bg=(AMBER if h in("nb_reclam","nb_non_clot") else DARK))
ws.row_dimensions[1].height = 30
for r,(idl,num,idp,typ,act) in enumerate(LIGNES,2):
    bg = stripe(r, even=LT_GREEN)
    cell(ws,r,1,idl,bold=True,color=GREEN,align='center',bg=bg,name='Consolas')
    cell(ws,r,2,num,align='center',bg=bg)
    cell(ws,r,3,idp,bold=True,color=PURPLE,align='center',bg=bg,name='Consolas')
    cell(ws,r,4,typ,align='center',bg=bg)
    ac = cell(ws,r,5,act,align='center',bg=bg,bold=True,
              color=(GREEN if act else RED))
    # helper formules (jointure réclamations)
    cell(ws,r,6,f'=COUNTIF(reclamations!$C:$C,$A{r})',align='center',bg=LT_AMBER,bold=True,color=AMBER)
    cell(ws,r,7,f'=COUNTIFS(reclamations!$C:$C,$A{r},reclamations!$H:$H,0)',align='center',bg=LT_AMBER,bold=True,color=AMBER)
for c,w in zip("ABCDEFG",[12,10,16,14,9,12,13]): ws.column_dimensions[c].width = w
box(ws,1,len(LIGNES)+1,1,7)
cell(ws, len(LIGNES)+3, 1,
     "Les colonnes ambre se calculent par jointure sur [reclamations] — elles alimentent le tableau de bord.",
     italic=True, color=SLATE, size=9)

# ────────────────────────────────────────────────────────────────────────────
#  3. passe_par  (table de jonction)
# ────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("passe_par"); setup(ws); ws.freeze_panes = "A3"
ws.merge_cells('A1:C1')
cell(ws,1,1,"⚡ TABLE DE JONCTION — une ligne dessert plusieurs communes (relation n–n)",
     italic=True,color=TEAL,bg=LT_TEAL,size=9)
for i,h in enumerate(["id_ligne","id_commune","nom_commune (lookup)"],1):
    header(ws,2,i,h,bg=TEAL)
ws.row_dimensions[2].height = 24
for r,(idl,idc) in enumerate(PASSE_PAR,3):
    bg = stripe(r, even=LT_TEAL)
    cell(ws,r,1,idl,bold=True,color=GREEN,align='center',bg=bg,name='Consolas')
    cell(ws,r,2,idc,bold=True,color=PURPLE,align='center',bg=bg,name='Consolas')
    cell(ws,r,3,f'=INDEX(ref_communes!$B:$B,MATCH($B{r},ref_communes!$A:$A,0))',
         color=SLATE,italic=True,bg=bg)
for c,w in zip("ABC",[12,14,28]): ws.column_dimensions[c].width = w
box(ws,2,len(PASSE_PAR)+2,1,3)
ex = len(PASSE_PAR)+4
cell(ws,ex,1,"Nb de lignes passant par Meudon (C003) :",bold=True,color=TEAL)
cell(ws,ex,2,'=COUNTIF($B:$B,"C003")',align='center',bold=True,color=TEAL,bg=LT_TEAL)

# ────────────────────────────────────────────────────────────────────────────
#  4. ref_communes
# ────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("ref_communes"); setup(ws); ws.freeze_panes = "A2"
for i,h in enumerate(["id_commune","nom_commune"],1): header(ws,1,i,h,bg=PURPLE)
ws.row_dimensions[1].height = 24
for r,(idc,nom) in enumerate(COMMUNES,2):
    bg = stripe(r, even=LT_PURPLE)
    cell(ws,r,1,idc,bold=True,color=PURPLE,align='center',bg=bg,name='Consolas')
    cell(ws,r,2,nom,bg=bg)
for c,w in zip("AB",[14,26]): ws.column_dimensions[c].width = w
box(ws,1,len(COMMUNES)+1,1,2)

# ────────────────────────────────────────────────────────────────────────────
#  5. ref_motifs
# ────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("ref_motifs"); setup(ws); ws.freeze_panes = "A2"
for i,h in enumerate(["id_motif","libelle_motif"],1): header(ws,1,i,h)
ws.row_dimensions[1].height = 24
for r,(idm,lib) in enumerate(MOTIFS,2):
    bg = stripe(r)
    cell(ws,r,1,idm,bold=True,color=DARK,align='center',bg=bg)
    cell(ws,r,2,lib,bg=bg)
for c,w in zip("AB",[12,26]): ws.column_dimensions[c].width = w
box(ws,1,len(MOTIFS)+1,1,2)

# ────────────────────────────────────────────────────────────────────────────
#  6. reclamations  (OLTP — table centrale)
# ────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("reclamations"); setup(ws); ws.freeze_panes = "A2"
cols = ["id_reclamation","date_reclamation","id_ligne","id_motif","id_commune",
        "objet_libre","date_cloture","cloturee","delai_traitement"]
for i,h in enumerate(cols,1):
    header(ws,1,i,h, bg=(AMBER if h=="delai_traitement" else BLUE))
ws.row_dimensions[1].height = 30
for r,(idr,d,idl,idm,idc,objet,dcl,clos) in enumerate(RECLAMATIONS,2):
    bg = stripe(r)
    cell(ws,r,1,idr,bold=True,color=DARK,align='center',bg=bg,name='Consolas')
    cell(ws,r,2,d,align='center',bg=bg,fmt='DD/MM/YYYY')
    cell(ws,r,3,idl,bold=True,color=GREEN,align='center',bg=bg,name='Consolas')
    cell(ws,r,4,idm,align='center',bg=bg)
    cell(ws,r,5,idc,bold=True,color=PURPLE,align='center',bg=bg,name='Consolas')
    cell(ws,r,6,objet,bg=bg)
    if dcl: cell(ws,r,7,dcl,align='center',bg=bg,fmt='DD/MM/YYYY')
    else:   cell(ws,r,7,None,bg=bg)
    if clos:
        cell(ws,r,8,1,align='center',bold=True,color=GREEN,bg=LT_GREEN)
    else:
        cell(ws,r,8,0,align='center',bold=True,color=RED,bg=LT_RED)
    cell(ws,r,9,f'=IF($H{r}=1,$G{r}-$B{r},"")',align='center',color=SLATE,bg=bg)
for c,w in zip("ABCDEFGHI",[16,16,11,10,12,52,15,11,16]): ws.column_dimensions[c].width = w
box(ws,1,len(RECLAMATIONS)+1,1,9)

# ════════════════════════════════════════════════════════════════════════════
#  7. 📊 TABLEAU DE BORD  (OLAP)
# ════════════════════════════════════════════════════════════════════════════
tb = wb.create_sheet("📊 Tableau de bord"); setup(tb)
widths = {'A':2,'B':24,'C':13,'D':14,'E':3,'F':2,'G':2,'H':26,'I':13,'J':3,'K':2,'L':2,'M':2}
for c,w in widths.items(): tb.column_dimensions[c].width = w
NR = len(RECLAMATIONS)

# Titre
tb.merge_cells('B1:I1')
cell(tb,1,2,"TABLEAU DE SUIVI DES RÉCLAMATIONS — 2024",bold=True,color=WHITE,bg=DARK,size=15,align='center')
tb.row_dimensions[1].height = 38
tb.merge_cells('B2:I2')
cell(tb,2,2,"Vue OLAP — se recalcule automatiquement à chaque saisie dans [reclamations]",
     italic=True,color=SLATE,align='center',size=9)

# ── KPI band (row 4 label / row 5 value) ──
kpis = [
    (2, "TOTAL",          "=COUNTA(reclamations!A:A)-1",            DARK,  None),
    (3, "CLÔTURÉES",      "=COUNTIF(reclamations!H:H,1)",           GREEN, None),
    (4, "NON CLÔTURÉES",  "=COUNTIF(reclamations!H:H,0)",           RED,   None),
    (8, "TAUX CLÔTURE",   "=C5/B5",                                 BLUE,  '0%'),
    (9, "DÉLAI MOYEN",    "=ROUND(AVERAGE(reclamations!I:I),1)",    AMBER, '0.0 "j"'),
]
for c_idx,label,formula,fg,fmt in kpis:
    cell(tb,4,c_idx,label,bold=True,color=fg,size=9,align='center')
    cell(tb,5,c_idx,formula,bold=True,color=fg,size=22,align='center',fmt=fmt)
tb.row_dimensions[5].height = 40

# ════════ LES 3 QUESTIONS (answer cards, rows 7-10) ════════
cell(tb,7,2,"LES 3 GRANDES QUESTIONS",bold=True,color=DARK,size=12)

# We place the detail tables LOWER and reference their ranges here.
# Ranges (computed below): presta D-col, commune I-col, ligne D-col.
P_HDR, P_DATA0 = 14, 15                 # PAR PRESTATAIRE header / first data row
P_N = len(PRESTATAIRES)
P_END = P_DATA0 + P_N - 1
L_HDR, L_DATA0 = 21, 22                  # PAR LIGNE
L_N = len(LIGNES)
L_END = L_DATA0 + L_N - 1
C_HDR, C_DATA0 = 14, 15                  # PAR COMMUNE (right block, col H/I)
C_N = len(COMMUNES)
C_END = C_DATA0 + C_N - 1

rng_presta_names = f"$B${P_DATA0}:$B${P_END}"
rng_presta_nclot = f"$D${P_DATA0}:$D${P_END}"
rng_ligne_ids    = f"$B${L_DATA0}:$B${L_END}"
rng_ligne_tot    = f"$D${L_DATA0}:$D${L_END}"
rng_comm_names   = f"$H${C_DATA0}:$H${C_END}"
rng_comm_tot     = f"$I${C_DATA0}:$I${C_END}"

cards = [
    # col, accent, light, emoji, titre, name_formula, val_formula, val_caption
    (2, RED, LT_RED, "🔴", "Prestataire le plus problématique",
        f'=INDEX({rng_presta_names},MATCH(MAX({rng_presta_nclot}),{rng_presta_nclot},0))',
        f'MAX({rng_presta_nclot})', "réclamations non clôturées"),
    (5, PURPLE, LT_PURPLE, "🟣", "Commune la plus impactée",
        f'=INDEX({rng_comm_names},MATCH(MAX({rng_comm_tot}),{rng_comm_tot},0))',
        f'MAX({rng_comm_tot})', "réclamations au total"),
    (8, GREEN, LT_GREEN, "🟢", "Ligne la plus problématique",
        f'=INDEX({rng_ligne_ids},MATCH(MAX({rng_ligne_tot}),{rng_ligne_tot},0))',
        f'MAX({rng_ligne_tot})', "réclamations au total"),
]
for col,accent,light,emoji,titre,fname,fval,cap in cards:
    tb.merge_cells(start_row=8,start_column=col,end_row=8,end_column=col+2)
    cell(tb,8,col,f"{emoji} {titre}",bold=True,color=WHITE,bg=accent,size=9,align='center',wrap=True)
    tb.row_dimensions[8].height = 28
    tb.merge_cells(start_row=9,start_column=col,end_row=9,end_column=col+2)
    cell(tb,9,col,fname,bold=True,color=accent,bg=light,size=16,align='center')
    tb.row_dimensions[9].height = 30
    tb.merge_cells(start_row=10,start_column=col,end_row=10,end_column=col+2)
    cc = tb.cell(row=10,column=col)
    cc.value = f'=TEXT({fval},"0")&"  {cap}"'
    cc.font = Font(italic=True,color=SLATE,size=9)
    cc.fill = PatternFill("solid",fgColor=light)
    cc.alignment = Alignment(horizontal='center')
    box(tb,8,10,col,col+2,side=Side(style='thin',color=accent))

# ════════ DÉTAIL — col gauche : PAR PRESTATAIRE puis PAR LIGNE ════════
cell(tb,13,2,"PAR PRESTATAIRE  (réclamations via jointure lignes_bus)",bold=True,color=RED,size=11)
for j,h in enumerate(["Prestataire","Total","Non clôt."],0):
    header(tb,P_HDR,2+j,h,bg=RED)
for k,(idp,nom,*_ ) in enumerate(PRESTATAIRES):
    r = P_DATA0+k; bg = stripe(r, even=LT_RED)
    cell(tb,r,2,nom,bg=bg)
    cell(tb,r,3,f'=SUMIF(lignes_bus!$C:$C,"{idp}",lignes_bus!$F:$F)',align='center',bold=True,bg=bg)
    cell(tb,r,4,f'=SUMIF(lignes_bus!$C:$C,"{idp}",lignes_bus!$G:$G)',align='center',bold=True,color=RED,bg=bg)
box(tb,P_HDR,P_END,2,4)

cell(tb,20,2,"PAR LIGNE  (COUNTIF direct sur la FK id_ligne)",bold=True,color=GREEN,size=11)
for j,h in enumerate(["Ligne","Prestataire","Total","Non clôt."],0):
    header(tb,L_HDR,2+j,h,bg=GREEN)
for k,(idl,num,idp,typ,act) in enumerate(LIGNES):
    r = L_DATA0+k; bg = stripe(r, even=LT_GREEN)
    cell(tb,r,2,idl,bold=True,color=GREEN,align='center',bg=bg,name='Consolas')
    # prestataire via double lookup
    cell(tb,r,3,f'=INDEX(prestataires!$B:$B,MATCH(INDEX(lignes_bus!$C:$C,MATCH($B{r},lignes_bus!$A:$A,0)),prestataires!$A:$A,0))',
         color=SLATE,bg=bg,size=9)
    cell(tb,r,4,f'=COUNTIF(reclamations!$C:$C,$B{r})',align='center',bold=True,bg=bg)
    cell(tb,r,5,f'=COUNTIFS(reclamations!$C:$C,$B{r},reclamations!$H:$H,0)',align='center',bold=True,color=RED,bg=bg)
box(tb,L_HDR,L_END,2,5)

# ════════ DÉTAIL — col droite : PAR COMMUNE, PAR MOTIF ════════
cell(tb,13,8,"PAR COMMUNE  (COUNTIF sur la FK id_commune)",bold=True,color=PURPLE,size=11)
for j,h in enumerate(["Commune","Total"],0):
    header(tb,C_HDR,8+j,h,bg=PURPLE)
for k,(idc,nom) in enumerate(COMMUNES):
    r = C_DATA0+k; bg = stripe(r, even=LT_PURPLE)
    cell(tb,r,8,nom,bg=bg)
    cell(tb,r,9,f'=COUNTIF(reclamations!$E:$E,"{idc}")',align='center',bold=True,color=PURPLE,bg=bg)
box(tb,C_HDR,C_END,8,9)

M_HDR = C_END+2      # PAR MOTIF header
M_DATA0 = M_HDR+1
cell(tb,M_HDR-1,8,"PAR MOTIF  (COUNTIF sur la FK id_motif)",bold=True,color=DARK,size=11)
for j,h in enumerate(["Motif","Total"],0):
    header(tb,M_HDR,8+j,h,bg=DARK)
for k,(idm,lib) in enumerate(MOTIFS):
    r = M_DATA0+k; bg = stripe(r)
    cell(tb,r,8,lib,bg=bg)
    cell(tb,r,9,f'=COUNTIF(reclamations!$D:$D,{idm})',align='center',bold=True,bg=bg)
M_END = M_DATA0+len(MOTIFS)-1
box(tb,M_HDR,M_END,8,9)

# ════════ PAR MOIS  (sous PAR LIGNE, col gauche) ════════
MO_HDR = L_END+2
MO_DATA0 = MO_HDR+1
cell(tb,MO_HDR-1,2,"PAR MOIS  (COUNTIFS sur une plage de dates)",bold=True,color=BLUE,size=11)
for j,h in enumerate(["Mois","Total"],0):
    header(tb,MO_HDR,2+j,h,bg=BLUE)
mois = [("Janvier 2024",1),("Février 2024",2),("Mars 2024",3),("Avril 2024",4)]
for k,(lib,m) in enumerate(mois):
    r = MO_DATA0+k; bg = stripe(r)
    cell(tb,r,2,lib,bg=bg)
    nxt = m+1
    cell(tb,r,3,
         f'=COUNTIFS(reclamations!$B:$B,">="&DATE(2024,{m},1),reclamations!$B:$B,"<"&DATE(2024,{nxt},1))',
         align='center',bold=True,color=BLUE,bg=bg)
MO_END = MO_DATA0+len(mois)-1
box(tb,MO_HDR,MO_END,2,3)
tb.merge_cells(start_row=2,start_column=2,end_row=2,end_column=9)

# Note bas de page
note_r = max(M_END, MO_END)+2
tb.merge_cells(start_row=note_r,start_column=2,end_row=note_r,end_column=9)
cell(tb,note_r,2,
     "💡 Aucune formule ci-dessus ne change quand on ajoute une réclamation : on saisit 1 ligne dans [reclamations], tout se recalcule. C'est la bascule OLTP → OLAP.",
     italic=True,color=SLATE,bg=LT_AMBER,wrap=True,size=9)
tb.row_dimensions[note_r].height = 30

# ── Conditional formatting : data bars sur les colonnes "Total" ──
tb.conditional_formatting.add(f"D{P_DATA0}:D{P_END}",
    DataBarRule(start_type='min',end_type='max',color="F8B4B4"))
tb.conditional_formatting.add(f"D{L_DATA0}:D{L_END}",
    DataBarRule(start_type='min',end_type='max',color="A7E8B8"))
tb.conditional_formatting.add(f"I{C_DATA0}:I{C_END}",
    DataBarRule(start_type='min',end_type='max',color="D6BBFB"))
tb.conditional_formatting.add(f"I{M_DATA0}:I{M_END}",
    DataBarRule(start_type='min',end_type='max',color="BBD3FB"))
# Highlight de la ligne max (le "coupable") en gras encadré
tb.conditional_formatting.add(f"B{L_DATA0}:E{L_END}",
    FormulaRule(formula=[f'$D{L_DATA0}=MAX({rng_ligne_tot})'], font=Font(bold=True), fill=PatternFill("solid",fgColor="FFE2A8")))
tb.conditional_formatting.add(f"H{C_DATA0}:I{C_END}",
    FormulaRule(formula=[f'$I{C_DATA0}=MAX({rng_comm_tot})'], font=Font(bold=True), fill=PatternFill("solid",fgColor="FFE2A8")))
tb.conditional_formatting.add(f"B{P_DATA0}:D{P_END}",
    FormulaRule(formula=[f'$D{P_DATA0}=MAX({rng_presta_nclot})'], font=Font(bold=True), fill=PatternFill("solid",fgColor="FFE2A8")))

# ════════════════════════════════════════════════════════════════════════════
#  Ordre des onglets + couleurs (OLTP / OLAP)
# ════════════════════════════════════════════════════════════════════════════
order = ["Lexique","prestataires","lignes_bus","passe_par","ref_communes",
         "ref_motifs","reclamations","📊 Tableau de bord"]
wb._sheets = [wb[s] for s in order]

tabcolors = {
    "Lexique":"1E3A5F",
    "prestataires":GREY, "lignes_bus":GREY, "passe_par":TEAL,
    "ref_communes":GREY, "ref_motifs":GREY,
    "reclamations":BLUE,          # OLTP central
    "📊 Tableau de bord":OLIVE,   # OLAP
}
for s,col in tabcolors.items():
    wb[s].sheet_properties.tabColor = col

out = "exercices/reclamations_complet.xlsx"
wb.save(out)
print(f"✅ {out}")
print(f"   {len(RECLAMATIONS)} réclamations · {len(order)} onglets")
print("   Réponses attendues : Transdev (8 non clôt.) · Meudon (7) · L169 (7)")
